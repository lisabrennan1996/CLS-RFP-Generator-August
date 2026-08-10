"""Daily batch job: pulls design/ops fields for EVERY study out of Fabric in one pass
and appends any study not already in the extract Excel file. Exists because live
per-user Fabric queries (fabric_design_lookup.py, the app's original approach) only
work for users who happen to have Fabric workspace access -- most don't. Running this
once, on a schedule, under an account that *does* have access, and having the app read
the resulting file instead (see fabric_extract_lookup.py) removes that dependency for
everyone else.

Incremental by design: existing rows are never rewritten or removed, only appended to --
each run reads the current file's "Study Alias" column, fetches the live set from
Fabric, and appends exactly the studies not already present. Re-running the same day (or
any day with no new studies) is a safe no-op (0 rows added).

Meant to be run on a schedule (Windows Task Scheduler) under an account with Fabric
workspace access to CDDA Analytics Reports_DEV.

Country Allocation is flattened into a single "Countries" cell, one country per
semicolon-pair-separated group, `Name|ABBR|STATUS` within a group -- reconstructed back
into the original list-of-dicts shape by fabric_extract_lookup.py.
"""
import argparse
import datetime as dt
import os
import re
from collections import defaultdict

import fabric_auth
import openpyxl

# Overridable via FABRIC_EXTRACT_PATH so the deployed webapp (no OneDrive client in the
# container) can point this at its own persistent-volume path instead -- see
# routers/admin.py's upload endpoint, which the desktop daily extract task pushes this
# same file to after writing its local OneDrive copy.
DEFAULT_OUTPUT_PATH = os.environ.get(
    "FABRIC_EXTRACT_PATH",
    r"C:\Users\L047081\OneDrive - Eli Lilly and Company\CLS AI Backend - Documents\RFP extract\fabric_study_extract.xlsx",
)

COLUMNS = [
    "Study Alias",
    "Phase",
    "Therapeutic Area",
    "Pediatric population?",
    "Countries",
    "Patients enrolled (randomized)",
    "Patients screened",
    "Protocol Approval (PA) date",
    "Planned First Patient Visit (FPV) date",
    "Planned Last Patient Visit (LPV) date",
    "FPET date",
    "LPET date",
    "Extracted On",
]

COUNTRY_GROUP_SEP = ";;"
COUNTRY_FIELD_SEP = "|"

PHASE_RE_ROMAN = {
    "i": "I", "ii": "II", "iii": "III", "iv": "IV",
    "1": "I", "2": "II", "3": "III", "4": "IV",
}


def _normalize_phase(raw):
    if not raw:
        return None
    m = re.search(r"phase\s*([ivx1-4]+)", raw, re.I)
    token = (m.group(1) if m else raw).strip().lower()
    return PHASE_RE_ROMAN.get(token)


def _yes_no(raw):
    if raw is None:
        return None
    v = str(raw).strip().lower()
    if v in ("y", "yes", "true", "1"):
        return "Yes"
    if v in ("n", "no", "false", "0"):
        return "No"
    return None


def _fetch_all_studies(conn):
    cur = conn.cursor()

    cur.execute(
        'SELECT [Study Alias], [Study Phase], [Therapeutic Area Description], '
        '[Will There Be Pediatric Participant] FROM [Study].[Study] WHERE [Study Alias] IS NOT NULL'
    )
    studies = {}
    for alias, phase_raw, ta, pediatric_raw in cur.fetchall():
        studies[alias] = {
            "Phase": _normalize_phase(phase_raw) or "",
            "Therapeutic Area": ta or "",
            "Pediatric population?": _yes_no(pediatric_raw) or "",
        }

    cur.execute(
        'SELECT [Study Alias], [Country Name], [Country Code], [Study Country Status] '
        'FROM [Study_Country].[Country] WHERE [Study Alias] IS NOT NULL AND [Country Name] IS NOT NULL'
    )
    countries_by_alias = defaultdict(list)
    for alias, name, abbr, status in cur.fetchall():
        countries_by_alias[alias].append(
            COUNTRY_FIELD_SEP.join([name or "", abbr or "", status or ""])
        )

    cur.execute(
        "SELECT [Study Alias], [Enrollment Type], [Total] FROM [Study].[Study Enrollment] "
        "WHERE [Study Alias] IS NOT NULL AND [Forecast, Planned, Actual] = 'Planned'"
    )
    for alias, enrollment_type, total in cur.fetchall():
        if alias not in studies or total is None:
            continue
        if enrollment_type == "Enrolled":
            studies[alias]["Patients enrolled (randomized)"] = total
        elif enrollment_type == "Screened":
            studies[alias]["Patients screened"] = total

    cur.execute(
        'SELECT [Study Alias], [Milestone Type], [Milestone Date] FROM [Study].[Study Milestone] '
        'WHERE [Study Alias] IS NOT NULL'
    )
    milestones_by_alias = defaultdict(dict)
    for alias, milestone_type, milestone_date in cur.fetchall():
        if milestone_type and milestone_date:
            milestones_by_alias[alias][milestone_type] = milestone_date

    for alias, row in studies.items():
        countries = countries_by_alias.get(alias)
        row["Countries"] = COUNTRY_GROUP_SEP.join(countries) if countries else ""

        ms = milestones_by_alias.get(alias, {})
        pa = ms.get("Protocol Approval")
        fpv = ms.get("First Patient Visit")
        lpv = ms.get("Primary Outcome Last Patient Visit") or ms.get("All Outcomes LPV")
        fpet = ms.get("First Patient Entered Treatment")
        lpet = ms.get("Last Patient Entered Treatment")
        row["Protocol Approval (PA) date"] = str(pa) if pa else ""
        row["Planned First Patient Visit (FPV) date"] = str(fpv) if fpv else ""
        row["Planned Last Patient Visit (LPV) date"] = str(lpv) if lpv else ""
        row["FPET date"] = str(fpet) if fpet else ""
        row["LPET date"] = str(lpet) if lpet else ""

    return studies


def _load_existing_aliases(path):
    if not os.path.exists(path):
        return None, set()
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    alias_col = header.index("Study Alias")
    existing = {
        row[alias_col]
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[alias_col]
    }
    return wb, existing


def main(output_path):
    app, cache = fabric_auth.msal_app()
    fabric_token = fabric_auth.get_token_blocking(app, cache, fabric_auth.FABRIC_SCOPE, "fabric")
    sql_token = fabric_auth.get_token_blocking(app, cache, fabric_auth.SQL_SCOPE, "sql")

    server, database = fabric_auth.resolve_sql_endpoint(fabric_token)
    conn = fabric_auth.sql_connect(server, database, sql_token)
    try:
        studies = _fetch_all_studies(conn)
    finally:
        conn.close()

    wb, existing_aliases = _load_existing_aliases(output_path)
    if wb is None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Studies"
        ws.append(COLUMNS)
    else:
        ws = wb.active

    today = dt.date.today().isoformat()
    added = 0
    for alias, row in studies.items():
        if alias in existing_aliases:
            continue
        ws.append([
            alias,
            row.get("Phase", ""),
            row.get("Therapeutic Area", ""),
            row.get("Pediatric population?", ""),
            row.get("Countries", ""),
            row.get("Patients enrolled (randomized)", ""),
            row.get("Patients screened", ""),
            row.get("Protocol Approval (PA) date", ""),
            row.get("Planned First Patient Visit (FPV) date", ""),
            row.get("Planned Last Patient Visit (LPV) date", ""),
            row.get("FPET date", ""),
            row.get("LPET date", ""),
            today,
        ])
        added += 1

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    fabric_auth.emit({
        "status": "ok",
        "added": added,
        "total_in_fabric": len(studies),
        "total_in_file": len(existing_aliases) + added,
        "output_path": output_path,
    })


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Daily incremental extract of all studies' design/ops fields from Fabric to Excel.")
    parser.add_argument("--output-path", default=DEFAULT_OUTPUT_PATH)
    args = parser.parse_args()
    try:
        main(args.output_path)
    except Exception as e:  # noqa: BLE001 -- single JSON-line contract, same as this app's other engine scripts
        fabric_auth.emit({"status": "error", "message": str(e)})
