"""Looks up one protocol alias's design/ops fields in the local Excel file
`fabric_daily_extract.py` maintains, instead of querying Fabric live. No MSAL, no
pyodbc, no network call at all -- this only ever touches a local file, so it works for
every user regardless of whether they have Fabric workspace access (the whole reason
this replaced the original live-query approach, fabric_design_lookup.py).

`lookup(protocol_alias, path)` is the importable entry point (used in-process by the
FastAPI webapp's `/api/fabric-design-fields` endpoint -- no subprocess needed there,
same reasoning as this engine's other in-process-import call sites). `main()` (the CLI
entry point the Tauri desktop app's Rust side still shells out to) is now a thin wrapper
that calls `lookup()` and prints its result as the one JSON line that caller expects --
same contract as before this refactor, no behavior change for that path:
  {"status": "not_found"}
  {"status": "error", "message": "..."}
  {"status": "ok", "fields": {<canonical field name>: <value>, ...}, "extracted_on": "..."}
"""
import argparse

import openpyxl

from fabric_daily_extract import COLUMNS, COUNTRY_FIELD_SEP, COUNTRY_GROUP_SEP, DEFAULT_OUTPUT_PATH


def _parse_countries(cell_value):
    if not cell_value:
        return []
    countries = []
    for group in str(cell_value).split(COUNTRY_GROUP_SEP):
        parts = group.split(COUNTRY_FIELD_SEP)
        name = parts[0] if len(parts) > 0 else ""
        if not name:
            continue
        countries.append({
            "name": name,
            "abbreviation": parts[1] if len(parts) > 1 and parts[1] else None,
            "status": parts[2] if len(parts) > 2 and parts[2] else None,
        })
    return countries


def _emit(payload):
    import json
    print(json.dumps(payload))


def lookup(protocol_alias, path=DEFAULT_OUTPUT_PATH):
    """Returns the same `{status, ...}` dict `main()` used to only print -- the
    importable form other callers (the FastAPI webapp) use directly, in-process."""
    import os

    if not os.path.exists(path):
        return {"status": "error", "message": f"extract file not found: {path}"}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {name: header.index(name) for name in COLUMNS if name in header}

    match = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col_idx["Study Alias"]] == protocol_alias:
            match = row
            break

    if match is None:
        return {"status": "not_found"}

    fields = {}
    phase = match[col_idx["Phase"]]
    if phase:
        fields["Phase"] = phase
    ta = match[col_idx["Therapeutic Area"]]
    if ta:
        fields["Therapeutic Area"] = ta
    pediatric = match[col_idx["Pediatric population?"]]
    if pediatric:
        fields["Pediatric population?"] = pediatric
    countries = _parse_countries(match[col_idx["Countries"]])
    if countries:
        fields["Country Allocation table"] = countries
        fields["Countries in scope"] = countries
    enrolled = match[col_idx["Patients enrolled (randomized)"]]
    if enrolled not in (None, ""):
        fields["Patients enrolled (randomized)"] = enrolled
    screened = match[col_idx["Patients screened"]]
    if screened not in (None, ""):
        fields["Patients screened"] = screened
    pa = match[col_idx["Protocol Approval (PA) date"]]
    fpv = match[col_idx["Planned First Patient Visit (FPV) date"]]
    lpv = match[col_idx["Planned Last Patient Visit (LPV) date"]]
    fpet = match[col_idx["FPET date"]]
    lpet = match[col_idx["LPET date"]]
    if pa:
        fields["Protocol Approval (PA) date"] = pa
    if fpv:
        fields["Planned First Patient Visit (FPV) date"] = fpv
    if lpv:
        fields["Planned Last Patient Visit (LPV) date"] = lpv
    if pa or fpv or lpv or fpet or lpet:
        fields["Trial Milestones (PA/FPV/LPV/FPET/LPET)"] = {
            "protocol_approval": pa or None,
            "fpv": fpv or None,
            "lpv": lpv or None,
            "fpet": fpet or None,
            "lpet": lpet or None,
        }

    return {
        "status": "ok",
        "fields": fields,
        "extracted_on": match[col_idx["Extracted On"]] if "Extracted On" in col_idx else None,
    }


def main(protocol_alias, path):
    _emit(lookup(protocol_alias, path))


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Look up design/ops fields for one protocol alias in the local Fabric extract file.")
    parser.add_argument("--protocol-alias", required=True)
    parser.add_argument("--extract-path", default=DEFAULT_OUTPUT_PATH)
    args = parser.parse_args()
    try:
        main(args.protocol_alias.strip(), args.extract_path)
    except Exception as e:  # noqa: BLE001 -- single JSON-line contract for the Rust caller
        _emit({"status": "error", "message": str(e)})
